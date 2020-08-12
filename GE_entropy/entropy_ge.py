# The script must be in the same folder as the *.xlsx files.

import shutil,glob
import os,sys # Working with file paths
from openpyxl import load_workbook #working with *.xlsx
import re #String matching name
import numpy as np
import matplotlib.pyplot as plt
import math #Standard Mathematical Functions
import pylab
from scipy import stats as st
from scipy import linalg as la


# Function _clasify_xlsx_file_names(1arg):
# 1arg: Directory Path
# return: Dictionary with *.xlxs file names. The keys are 'PC', 'sample'.
def _clasify_xlsx_file_names(dir_path):
    names ={}
    names['PC'] =[]
    names['sample'] =[]
    for f_name in os.listdir(dir_path):
        if f_name.startswith('PC') and f_name.endswith('xlsx'):
            names['PC'].append(f_name)
        if f_name.startswith('sample') and f_name.endswith('xlsx'):
            names['sample'].append(f_name)
    return names


# Function _path(1arg):
# 1arg: list of *.xlsx file names
# return: list of paths corresponding to *.xlsx file names.
def _path(names):        
    xls_loc=[]
    for name in names:
        xls_loc.append(os.path.abspath(name))
    return xls_loc


# Function _files_path(1arg*, 2arg*, 3arg):
# 1arg*: Function _clasify_xlsx_file_names
# 2arg*: Function _path
# 3arg: Directory Path   
# return: Dictionary with paths for *.xlxs file names. The keys are 'PC', 'sample'.
def _files_path(_clasify_xlsx_file_names, _path, dir_path):
    files_path={}
    list_of_names = _clasify_xlsx_file_names(dir_path)
    for key in list_of_names.keys():
        files_path[key] = _path(list_of_names[key])
    return files_path


# Function _copy_row_minus_1_from_xlsx(1arg, 2arg, 3arg):
# 1arg: Integer
# 2arg: Integer
# 3arg: .xlsx file
# return: List
def _copy_row_minus_1_from_xlsx(i, max_column_PC, PC_data_sheet):
    index = i-1
    vector=[]
    for k in range(1,max_column_PC +1):
        try:
            vector.append(float(PC_data_sheet.cell(row=index,column=k).value))
        except:
            print(PC_data_sheet.cell(row=index,column=k).value)
    return vector


# Function _copy_data_filter_by_stage(1arg):
# 1arg: Directory Path
# return: Dictionary of Dictionary of matrices
#       , List of missmatched files(e.g each 'sample' needs a 'PC').
def _copy_data_filter_by_stage(files_path):
    missmatched_files=[]
    missing_correspondence_test = False
    
    copy={}
    unclass_copy={}
    for PC_names in files_path['PC']:
        pc = re.search('/PC_(.*).xlsx', PC_names)
        for sample_names in files_path['sample']:
            sample = re.search('/sample_(.*).xlsx', sample_names)
            if pc.group(1) == sample.group(1):
                copy[pc.group(1)]={} #dictionary saving copy of normal and diseased data
                copy[pc.group(1)]['normal']=[]
                copy[pc.group(1)]['tumor']=[]
                
                unclass_copy[pc.group(1)]={}
                unclass_copy[pc.group(1)]['normal']=[]
                unclass_copy[pc.group(1)]['tumor']=[]
                
                wb1, wb2 = load_workbook(sample_names), load_workbook(PC_names)
                index_sheet, PC_data_sheet = wb1.active, wb2.active
                
                max_row, max_column = index_sheet.max_row, index_sheet.max_column
                
                max_column_PC = PC_data_sheet.max_column
                
                for j in range(1, max_column+1):#Looking for the column titled 'Sample Type'
                    if index_sheet.cell(row=1,column=j).value =='Sample Type':
                        for i in range(2,max_row+1):#Looking for 'Solid Tissue Normal' in rows
                            if index_sheet.cell(row=i,column=j).value =='Solid Tissue Normal':
                                row = \
                                _copy_row_minus_1_from_xlsx(i, max_column_PC, PC_data_sheet)
                                copy[pc.group(1)]['normal'].append(row)
                                unclass_copy[pc.group(1)]['normal'].append(row)
                                unclass_copy[pc.group(1)]['tumor'].append(row)
                            else:
                                row = \
                                _copy_row_minus_1_from_xlsx(i, max_column_PC, PC_data_sheet)
                                copy[pc.group(1)]['tumor'].append(row)
                                unclass_copy[pc.group(1)]['normal'].append(row)
                                unclass_copy[pc.group(1)]['tumor'].append(row)
            if pc.group(1) == sample.group(1):
                missing_correspondence_test = False
                break
            else:
                missing_correspondence_test = True
        if missing_correspondence_test == True:
            missmatched_files.append(pc.group(1))
                
    return copy,unclass_copy, missmatched_files


# Function _transpose_data(1arg):
# 1arg: Dictionary of Dictionary of matrices.
# return: Dictionary of Dictionary of transposed_matrices.
def _transpose_data(data):
    core_matrix = {}
    for label in data.keys():
        core_matrix[label] = {}
        for stage, core_matrix_T in data[label].items():
            core_matrix[label][stage] = np.around(np.array(core_matrix_T, dtype=np.float64).T, decimals = 10)
            
    return core_matrix


def _low_dimension_core_matrix(core_matrix, LOW_DIMENSION):
        unif_restricted_core_matrix={}
        for label in core_matrix.keys():
            unif_restricted_core_matrix[label]={}
            for stage in core_matrix[label].keys():
                unif_restricted_core_matrix[label][stage]=[]
                for i in range(LOW_DIMENSION):
                    row=core_matrix[label][stage][i]
                    unif_restricted_core_matrix[label][stage].append(row)
        return unif_restricted_core_matrix
    
    
# Function _complexity_map(1arg, 2arg):
# 1arg: Dictionary of Dictionary of floats.
# 2arg: Dictionary of Dictionary of floats.
# return: Plot
def _complexity_map(mean, covariance_matrix ,DIMENSION, EXCLUSION_LIST):  
    entropy_dict = _H_multiv_gaussian(covariance_matrix,DIMENSION)
    
    delta_H = _delta_H(entropy_dict)
    overlap, log_overlap = _overlap(mean, covariance_matrix ,DIMENSION)
    
    x1,x2,x3,y,z= [],[],[],[],[]
    
    for label in log_overlap.keys():
        if (label in EXCLUSION_LIST):
            pass
        else:
            y.append(log_overlap[label])
            x1.append(entropy_dict[label]['normal'])
            x2.append(entropy_dict[label]['tumor'])
            x3.append(delta_H[label])
            z.append(label)
        
    
    fig, ax = plt.subplots()
    ax.scatter(x3, y, label=r'$\Delta S=S_{tumor}-S_{normal}$')
    ax.scatter(x1, y, label=r'$S_{normal}$')
    ax.scatter(x2, y, label=r'$S_{tumor}$')
    
    for i, txt in enumerate(z):
        ax.annotate(txt, (x1[i], y[i]))
        ax.annotate(txt, (x2[i], y[i]))
        ax.annotate(txt, (x3[i], y[i]))
    
    ident = [0.0, 100.0]
    fit= np.polyfit(x3,y,1)#Polynomial coefficients, highest power first
    
    fitted_y =[fit[0]*i+fit[1] for i in ident] 
    
    
    plt.plot(ident,fitted_y,'b--',label=r"$m \approx {0},n \approx {1}$".format(np.around(fit[0], 2),np.around(fit[1], 2)),linewidth=1)
    #plt.plot(ident,ident,'k--',label=r'$\alpha =1,\beta =0$',linewidth=1)
    plt.xlabel(r"$\Delta S, S$ (nats)")
    plt.ylim(0, 100)
    plt.ylabel(r"$-\ln \,I$")
    plt.title("Overlap $vs$ Entropy")
    pylab.legend(loc='lower left')
    plt.savefig('overlapentropy_fig.pdf')   
    
    return fit
 


# Function _PC_pair_plot(1arg, 2arg, 3arg):
# 1arg: Dictionary of Dictionary of matrices.
# 2arg: Integer. PC_i, axis x.
# 3arg: Integer. PC_j, axis y.
# return: Plot
def _PC_pair_plot(core_matrix, i, j,location_name):
    #location_name = 'LUSC' # 'LUSC' shows a well defined separation between 'normal' and 'tumor'
    PC_j={}
    PC_i={}
    names = []
    for label in core_matrix.keys():
        names.append(label)
        PC_i[label]={}
        PC_j[label]={}
        for stage in core_matrix[label].keys():
            PC_i[label][stage]=core_matrix[label][stage][i-1]
            PC_j[label][stage]=core_matrix[label][stage][j-1]
        
    fig, ax = plt.subplots()
    ax.scatter(PC_i[location_name]['normal'], PC_j[location_name]['normal'], label='normal')
    ax.scatter(PC_i[location_name]['tumor'],PC_j[location_name]['tumor'], label='tumor')
    
    #for i, txt in enumerate(names):
        #ax.annotate(txt, (PC_i[location_name]['normal'][i], PC_j[location_name]['normal'][i]))
        #ax.annotate(txt, (PC_i[location_name]['tumor'][i], PC_j[location_name]['tumor'][i]))
        
    plt.xlabel("PC_{0}".format(i))
    plt.ylabel("PC_{0}".format(j))
    plt.title("PC_map_{0}".format(location_name))
    pylab.legend(loc='lower left')
    plt.savefig('pairplot_fig.pdf')   
    return None




# Function _std_dict(1arg):
# 1arg: Dictionary of Dictionary of matrices.
# return: Dictionary of Dictionary of vectors containing the standard deviation
#         of each matrix row.
def _std_dict(core_matrix):
    std={}
    for label in core_matrix.keys():
        std[label]={}
        for stage in core_matrix[label].keys():
            std[label][stage]=[]
            for row in core_matrix[label][stage]:
                std[label][stage].append(np.std(row))
    return std


# Function _mean_dict(1arg):
# 1arg: Dictionary of Dictionary of matrices.
# return: Dictionary of Dictionary of vectors containing the mean of each matrix row.
def _mean_dict(core_matrix):
    mean={}
    for label in core_matrix.keys():
        mean[label]={}
        for stage in core_matrix[label].keys():
            mean[label][stage]=[]
            for row in core_matrix[label][stage]:
                mean[label][stage].append(np.mean(row))
    return mean


def _covariance_matrix_(core_matrix):
    covariance_matrix={}
    for label in core_matrix.keys():
        covariance_matrix[label]={}
        for stage in core_matrix[label].keys():
            covariance_matrix[label][stage]=\
            np.cov(core_matrix[label][stage], ddof=1 )#np.around(, decimals = 10)
    
    return covariance_matrix

#Returns Diagonal_values, Regular
def _matrix_SVD(covariance_matrix):
    s, cov_inv = {},{}
    for label in covariance_matrix.keys():
        s[label], cov_inv[label] = {},{}
        for stage in covariance_matrix[label].keys():
            u, ss, vh =np.linalg.svd(covariance_matrix[label][stage])
            s[label][stage]=np.diag(ss)
            cov_inv[label][stage]=np.matmul(np.matmul(vh.T,np.linalg.inv(s[label][stage])),u.T)
    return s, cov_inv

# ENTROPY of a MULTIVARIATE GAUSSIAN
def _H_multiv_gaussian(covariance_matrix,DIMENSION):
    H_={}
    S_, cov_inv = _matrix_SVD(covariance_matrix)
    for label in covariance_matrix.keys():
        H_[label]={}
        for stage in covariance_matrix[label].keys():
            sum_log = sum(math.log(S_[label][stage][i][i]) for i in range(DIMENSION))
            H_[label][stage] =\
            0.5*(sum_log + DIMENSION*(1 + math.log(2*math.pi) ))   
    return H_

def _delta_H(entropy_dict):
    delta_H={}
    for label in entropy_dict.keys():
        delta_H[label]= entropy_dict[label]['tumor']\
                       -entropy_dict[label]['normal']
    return delta_H                   


# DIVERGENCE between two MULTIVARIATE GAUSSIANS
def _D_multiv_gaussians(std, mean, stage, ref_stage, DIMENSION):
    D_={}
    
    for label in std.keys():
        m_sum = sum(((mean[label][stage][i]-mean[label][ref_stage][i])**2)/(std[label][ref_stage][i]**2) for i in range(DIMENSION))
        sum_log = sum(math.log(math.pow( math.sqrt(2 * math.pi) * std[label][ref_stage][i], std[label][ref_stage][i]/std[label][stage][i]))  for i in range(DIMENSION))
        D_[label] = DIMENSION/2 + m_sum + sum_log
    
    return D_


def _overlap(mean, covariance_matrix, DIMENSION):
    overlap, log_overlap={},{}
    S, Cov_Inv=_matrix_SVD(covariance_matrix)
    A, B = DIMENSION/2 , DIMENSION/4
    for label in mean.keys():
        inv_Vc = Cov_Inv[label]['normal']+Cov_Inv[label]['tumor']
        Vc = np.linalg.inv(inv_Vc)
        
        u_Vc, s_Vc, vh_Vc = np.linalg.svd(Vc)
        u_inv_Vc, s_inv_Vc, vh_inv_Vc = np.linalg.svd(inv_Vc)
        det_Vc, det_inv_Vc = np.prod(s_Vc), np.prod(s_inv_Vc)
        
        u_inv_Vn, s_inv_Vn, vh_inv_Vn = np.linalg.svd(Cov_Inv[label]['normal'])
        u_inv_Ve, s_inv_Ve, vh_inv_Ve = np.linalg.svd(Cov_Inv[label]['tumor'])
        
        
        log_sum = sum( -math.log(s_inv_Vn[i]) - math.log(s_inv_Ve[i]) +math.log(s_inv_Vc[i])  for i in range(DIMENSION))
        
        
        eta_n = np.matmul(Cov_Inv[label]['normal'],mean[label]['normal'])
        eta_e = np.matmul(Cov_Inv[label]['tumor'],mean[label]['tumor'])
        eta_c = eta_n + eta_e
        
        mixed = np.matmul(np.transpose(eta_n),np.matmul(covariance_matrix[label]['normal'],eta_n))\
               +np.matmul(np.transpose(eta_e),np.matmul(covariance_matrix[label]['tumor'],eta_e))\
               -np.matmul(np.transpose(eta_c),np.matmul(Vc,eta_c))\
        
        arg= -0.5*(DIMENSION*math.log(2*math.pi) + log_sum + mixed)
        
        
        overlap[label]= math.pow(2,A)*math.pow(2*math.pi,B)*math.exp(arg)*math.pow(det_inv_Vc,-0.25)
        #log_overlap[label]=-math.log(overlap[label])
        log_overlap[label]=-A*math.log(2)-B*math.log(2*math.pi)-arg +0.25*math.log(det_inv_Vc)
    return overlap, log_overlap

def _Div_map(std, mean, DIMENSION,EXCLUSION_LIST):  
    div_n_t = _D_multiv_gaussians(std, mean, 'normal', 'tumor', DIMENSION)
    div_t_n = _D_multiv_gaussians(std, mean, 'tumor', 'normal', DIMENSION)
    
    x,y,z= [],[],[]
    
    for label in div_t_n.keys():
        if (label in EXCLUSION_LIST):
            pass
        else:
            y.append(div_n_t[label])
            x.append(div_t_n[label])
            z.append(label)
        
    fig, ax = plt.subplots()
    ax.scatter(x, y)
    
    for i, txt in enumerate(z):
        ax.annotate(txt, (x[i], y[i]))
    ident = [0.0, 300.0]
    plt.plot(ident,ident)
    
    plt.xlabel("D(tumor||normal)_nats")
    plt.ylabel("D(normal||tumor)_nats")
    plt.title("Divergence_map_DIMENSION_{0}".format(DIMENSION))
        
    return None

def _norm_plot(covariance_matrix, entropy_dict):
    norm_dict={}
    for label in covariance_matrix.keys():
        norm_dict[label]={}
        for stage in covariance_matrix[label].keys():
            norm_dict[label][stage] = math.log(np.linalg.norm(covariance_matrix[label][stage]))
    
    x1, y1, x2, y2, z= [], [], [], [], []
    
    for label in norm_dict.keys():
        if (label in EXCLUSION_LIST):
            pass
        else:
            y1.append(norm_dict[label]['normal'])
            x1.append(entropy_dict[label]['normal'])
            y2.append(norm_dict[label]['tumor'])
            x2.append(entropy_dict[label]['tumor'])
            z.append(label)
        
    fig, ax = plt.subplots()
    ax.scatter(x1, y1)
    ax.scatter(x2, y2)
    
    for i, txt in enumerate(z):
        ax.annotate(txt, (x1[i], y1[i]))
        ax.annotate(txt, (x2[i], y2[i]))
    
    plt.xlabel("S_nats")
    plt.ylabel("norm")
    plt.title("Temperature_map_DIMENSION_{0}".format(20))
    return None 

# The main program
if __name__ == "__main__":

    src='../PCA_tcga/'
    dst='.'
    files = glob.iglob(os.path.join(src, "*.xlsx"))
    for file in files:
        if os.path.isfile(file):
            shutil.copy2(file, dst)

    EXCLUSION_LIST =['READ','ESCA','CESC', 'Adren', 'GBM', 'PAAD', 'BLCA']# BECAUSE FEW 'normal' SAMPLES, OTHERWISE EMPTY.
    dir_path = os.path.abspath('.') #Set absolute path to directory PCA_tcga
    files_path = _files_path(_clasify_xlsx_file_names, _path, dir_path)
    data, unclass_data, missmatched_files = _copy_data_filter_by_stage(files_path)
    core_matrix = _transpose_data(data)
    unclass_core_matrix= _transpose_data(unclass_data)    

    #############################
    std = _std_dict(core_matrix)
    mean = _mean_dict(core_matrix)
    covariance_matrix = _covariance_matrix_(core_matrix)
    
    #divergence_n_t=_D_multiv_gaussians(std, mean, stage='normal', ref_stage='tumor', DIMENSION=20)
    #divergence_t_n=_D_multiv_gaussians(std, mean, stage='tumor', ref_stage='normal', DIMENSION=20)
    ############################
    unclass_std = _std_dict(unclass_core_matrix)
    unclass_mean = _mean_dict(unclass_core_matrix)
    unclass_covariance_matrix = _covariance_matrix_(unclass_core_matrix)
    unclass_H_multiv_gaussian=_H_multiv_gaussian(unclass_covariance_matrix,DIMENSION=20)

    #for i in range(2,20):
        #_complexity_map(mean, unclass_mean, covariance_matrix ,unclass_covariance_matrix ,DIMENSION=i)
        #_Div_map(std, mean, DIMENSION=i)
    
    
    low_DIMENSION_core_matrix=_low_dimension_core_matrix(core_matrix, LOW_DIMENSION=20)
    low_DIMENSION_covariance_matrix = _covariance_matrix_(low_DIMENSION_core_matrix)
    low_DIMENSION_mean = _mean_dict(low_DIMENSION_core_matrix)
    low_DIMENSION_std = _std_dict(low_DIMENSION_core_matrix)
    linear_fit =_complexity_map(low_DIMENSION_mean,  low_DIMENSION_covariance_matrix ,20, EXCLUSION_LIST)#Polynomial coefficients, highest power first
    _Div_map(low_DIMENSION_std, low_DIMENSION_mean,20, EXCLUSION_LIST)
    
    
    overlap, log_overlap = _overlap(mean, covariance_matrix, 20)
    
    entropy_dict = _H_multiv_gaussian(covariance_matrix,DIMENSION=20)
    
    delta_H = _delta_H(entropy_dict)
    _PC_pair_plot(core_matrix, 1, 2,'LUSC') 
    _norm_plot(covariance_matrix, entropy_dict)
    
    def _partition(covariance_matrix, DIMENSION):
        partition={}
        for label in covariance_matrix.keys():
            partition[label]={}
            for stage in covariance_matrix[label].keys():
                v,s,u = np.linalg.svd(covariance_matrix[label][stage])
                det = np.prod(s)
                partition[label][stage] = pow(2*math.pi, DIMENSION*0.5)*math.sqrt(det)
        return partition
    partition=_partition(covariance_matrix,20)    

    files = glob.iglob(os.path.join(".", "*.xlsx"))
    for file in files:
        if os.path.isfile(file):
            os.remove(file)
